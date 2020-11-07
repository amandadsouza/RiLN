import sqlite3 
import nltk as nltk 
from operator import itemgetter 
import constantes 
import json 
import csv 
from openpyxl import Workbook 
import preProcessamentoTextual as pre 

class BD:

    def __init__(self, nameDB):
        self.nameDB = nameDB
        self.conn = None
        self.cursor = None

    def __enter__(self):
        self.conn = sqlite3.connect(self.nameDB)
        self.cursor = self.conn.cursor() 

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.conn.commit()
        self.conn.close() 

    def criarBancoDeDados(self):
        """[summary]
        """        
        self.cursor.execute("""CREATE TABLE if not exists respostas 
                            (   cd_documento integer,
                                ds_documento text,
                                cd_editor_registro integer,
                                tp_objeto text,
                                cd_metadado integer,
                                cd_campo integer,
                                ds_campo text,
                                resposta text,
                                respostaTratada text,
                                quantStopWords integer,
                                quantNoStopWords integer,
                                bigramas text,
                                quantBigramas integer,
                                trigramas text,
                                quantTrigramas integer,
                                quantCID integer
                                 )
                            """ )

    def listarRespostas(self, criterio):
        """ Seleciona os registros da tabela resposta de acordo com criterio
        
        Arguments:
            criterio {str} -- Se devem ser TODAS, ANAMNESE ou EVOLUCAO 

        Returns:
            list -- Um array com os registros retornados do SELECT 
        """        
        selecao =   """
                    SELECT   rowid
                            , cd_documento 
                            , ds_documento 
                            , cd_editor_registro 
                            , tp_objeto 
                            , cd_metadado 
                            , cd_campo 
                            , ds_campo 
                            , resposta 
                            , respostaTratada 
                            , quantStopWords 
                            , quantNoStopWords  
                            , bigramas 
                            , quantBigramas 
                    FROM respostas
                    """

        if criterio == "TODAS":
            dataset = self.cursor.execute(selecao).fetchall()
        else: 
            selecao = selecao + " WHERE ds_documento like ?"
            dataset = self.cursor.execute(selecao, (criterio+'%',)).fetchall()
        return dataset

    def exportarBancoDeDados(self):
        data = self.listarRespostas("TODAS")
        wb = Workbook()
        ws0 = wb.active
        ws0.title = 'BancoDeDados'
        conjunto = []
        for linha in data:
            conjunto.clear()
            for coluna in linha:
                conjunto.append(coluna)
            ws0.append(conjunto) 
        arq = '/Users/eduardofelipe/Google Drive/Doutorado-Amanda/Resultados/bancoDeDados.xlsx'
        wb.save(arq)

    def gravarNovoTexto(self, textoTratado, id):
        """[summary]
        
        Arguments:
            textoTratado {[type]} -- [description]
            id {[type]} -- [description]
        """        
        self.cursor.execute("""UPDATE respostas SET respostaTratada = ? WHERE rowid = ?""", (textoTratado, id,))

    def gravarQuantStopWords(self, id, estrutura):
        """Grava a quantidade de stopwords identificada neste registro especifico
        
        Arguments:
            id {int} -- Chave primaria
            estrutura {list} -- Quantidade de stopwords na primeira posicao e quantidade de tokens que nao sao stopwords na segunda posicao
        """        
        self.cursor.execute("""UPDATE respostas SET quantStopWords = ?, quantNoStopWords  = ? WHERE rowid = ? """,(estrutura['stop'], estrutura['noStop'], id, )) 

    def gravarRegistroComQuantCID(self, id, quant):
        """Grava a quantidade de codigos CID naquele registro especifico

        Arguments:
            id {int} -- Chave primaria 
            quant {int} -- Quantidade de codigos CID encontrados neste registro
        """        
        self.cursor.execute("""UPDATE respostas SET quantCID = ? WHERE rowid = ? """,(quant, id, )) 

    def gravarBigramasNoBD(self, id, bigramas, tamanho):
        """ Recebe os dados do processamento de bigramas e grava na tabela

        Args:
            id (int): Codigo da linha (chave)
            bigramas (str): Quais sao os bigramas
            tamanho (int): Quantidade de bigramas
        """        
        self.cursor.execute("""UPDATE respostas SET bigramas = ?, quantBigramas = ? WHERE rowid = ? """,(bigramas, tamanho, id, ))

    def gravarTrigramasNoBD(self, id, trigramas, tamanho):
        """ Recebe os dados do processamento de trigramas e grava na tabela

        Args:
            id (int): Codigo da linha (chave)
            trigramas (str): Quais sao os trigramas
            tamanho (int): Quantidade de trigramas
        """        
        self.cursor.execute("""UPDATE respostas SET trigramas = ?, quantBigramas = ? WHERE rowid = ? """,(trigramas, tamanho, id, ))

    def trataStopWords(self, gravar):
        """ Percorre toda a tabela e identifica cada token, para contar se eh stopword ou nao 
            ja grava as quantidades em cada registro no banco de dados.
            Deve se chamada uma vez com S para gravar os dados no proprio banco. 
            Depois pode-se chamar com N apenas para gerar a saida em XLS
        Args:
            gravar (str): 'S' ou 'N'
        """        
        wb = Workbook() 
        ws1 = wb.active 
        ws1.title = 'quantStopWords'
        arq = constantes.PATH_RESULTADOS + 'QuantidadeStopWords.xlsx'
        dados = self.listarRespostas('TODAS') 
        for linha in dados:
            if (linha[8] is not None): 
                novoTexto = pre.limpaTexto(linha[8]) 
                resposta  = pre.contarStopWords(novoTexto) 
                if (gravar == 'S'):
                    self.gravarNovoTexto(novoTexto, linha[0])
                    self.gravarQuantStopWords(linha[0], resposta)
                ws1.append([novoTexto, resposta['stop'], resposta['noStop']]) 
        wb.save(arq)

    def separaStopWords(self):
        """ A partir de um tipo de pesquisa, identifica, separa e conta as stopWords 
            Grava a saida em arquivo excel 
        """ 
        colecao = {}
        wb = Workbook()
        ws0 = wb.active
        ws0.title = 'stopWords-TODAS'
        ws1 = wb.create_sheet(title='stopWords-ANAMNESE') 
        ws2 = wb.create_sheet(title='stopWords-EVOLUCAO') 
        arq = constantes.PATH_RESULTADOS + 'stopWords-.xlsx' 
        arqTodas = constantes.PATH_RESULTADOS + 'stopWords-ParaNuvem-Todas.txt'
        strTodas = ''
        arqAnamnese = constantes.PATH_RESULTADOS + 'stopWords-ParaNuvem-Anamnese.txt'
        strAnamnese = ''
        arqEvolucao = constantes.PATH_RESULTADOS + 'stopWords-ParaNuvem-Evolucao.txt'
        strEvolucao = ''
        tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO'] 
        for tipo in tipoResposta:
            tabela = self.listarRespostas(tipo)
            for linha in tabela:
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    for token in tokens:
                        if (token in pre.stopWords):
                            valor = colecao.get(token)
                            if valor is None:
                                valor = 0
                            colecao[token] = valor + 1
            for item in colecao:
                quant =  colecao.get(item)
                if tipoResposta.index(tipo) == 0:
                    ws0.append([item, quant])
                    strTodas += ' '+ pre.repeteString(item, quant)
                elif tipoResposta.index(tipo) == 1:
                    ws1.append([item, quant])
                    strAnamnese += ' '+ pre.repeteString(item, quant)
                elif tipoResposta.index(tipo) == 2:
                    ws2.append([item, quant])
                    strEvolucao += ' '+ pre.repeteString(item, quant)
            
            colecao.clear()
            tabela.clear()
        wb.save(arq)
        pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strTodas)).strip(), arqTodas)
        pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strAnamnese)).strip(), arqAnamnese)
        pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strEvolucao)).strip(), arqEvolucao)

    # def trataNgrams(self):
    #     """[summary]
    #     """        
    #     dados = self.listarRespostas()
    #     for linha in dados:
    #         print(linha[9])
    #         if (linha[9] is not None):
    #             fraseTokenizada = nltk.word_tokenize(linha[9])
    #             bigrama = nltk.bigrams(fraseTokenizada)
    #             print(bigrama)
    #         exit()


    def countCIDForArq(self):
        """ Percorre a tabela dependendo do criterio e identifica quais os codigos CID estao presentes em cada registro, 
            realiza uma soma da presenca dos codigos para uma visao estatistica dos mais usados de acordo com o filtro do criterio
            gera um arquivo txt como resultado e grava na pasta RESULTADOS (definida no arquivo de constantes) 
            dois arquivos sao gerados como saida .txt e .csv 
        """ 
        colecao = {} 
        tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO']
        cid = pre.carregarArquivoComoArray(constantes.ARQ_CID)
        for tipo in tipoResposta:
            tabela = self.listarRespostas(tipo) 
            for linha in tabela:
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    for token in tokens:
                        if token in cid:
                            if token in colecao:
                                valor = colecao.get(token)
                                colecao[token] = valor + 1
                            else:
                                colecao[token] = 1
            colecaoOrdenada = ""
            colecaoOrdenada = OrderedDict(sorted(colecao.items(), key=itemgetter(1)))
            with open(constantes.PATH_RESULTADOS+'cid-'+tipo+'.csv', 'w') as f:
                writer = csv.writer(f)
                for chave, valor in colecaoOrdenada.items():
                    writer.writerow([chave, valor])
            arquivo = open(constantes.PATH_RESULTADOS+'cid-'+tipo+'.txt', 'w')
            json.dump(colecaoOrdenada, arquivo, ensure_ascii=False)
            colecao.clear()
            arquivo.close()

    def countCIDForBD(self):
        """Percorre toda a tabela e identifica quais os codigos CID estao presentes em cada registro, 
            e realiza uma soma da presenca dos codigos para uma visao unica de cada tupla
            grava o resultado em um campo (coluna) no banco de dados quantCID 
        """        
        tabela = self.listarRespostas('TODAS')
        cid = pre.carregarArquivoComoArray(constantes.ARQ_CID)
        for linha in tabela:
            quantCid = 0
            if (linha[9]):
                tokens = pre.tokenizeString(linha[9])
                for token in tokens:
                    if token in cid:
                        quantCid += 1 
                self.gravarRegistroComQuantCID(linha[0], quantCid)

    def countSiglaForTxt(self):
        """ [summary]
        """        
        colecao = {} 
        tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO']
        siglas = pre.carregarArquivoComoArray(constantes.ARQ_SIGLAS)
        for tipo in tipoResposta:
            tabela = self.listarRespostas(tipo) 
            for linha in tabela:
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    for token in tokens:
                        if token in siglas:
                            if token in colecao:
                                valor = colecao.get(token)
                                colecao[token] = valor + 1
                            else:
                                colecao[token] = 1
            colecaoOrdenada = ""
            colecaoOrdenada = OrderedDict(sorted(colecao.items(), key=itemgetter(1)))
            with open(constantes.PATH_RESULTADOS+'siglas-'+tipo+'.csv', 'w') as f:
                writer = csv.writer(f)
                for chave, valor in colecaoOrdenada.items():
                    writer.writerow([chave, valor])            
            arquivo = open(constantes.PATH_RESULTADOS+'siglas-'+tipo+'.txt', 'w')
            json.dump(colecaoOrdenada, arquivo, ensure_ascii=False, indent=4)
            colecao.clear()
            arquivo.close()

    def processaBigramas(self):
        """ Identifica e conta os bigramas. 
            Segmenta por Anamnese, Evolucao e Todos. 
        """        
        # processa toda a tabela 
        wb = Workbook()
        ws0 = wb.active
        ws0.title = 'bigramas-TODAS'
        ws1 = wb.create_sheet(title='bigramas-ANAMNESE')
        ws2 = wb.create_sheet(title='bigramas-EVOLUCAO')
        arq = 'bigramas.xlsx' 
        listaTodosBigramas = []
        tipoResposta = ['ANAMNESE', 'EVOLUCAO']
        for tipo in tipoResposta:
            tabela = self.listarRespostas(tipo)
            for linha in tabela:
                #rowId = linha[0]
                #texto = ''
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    bigramas = list(nltk.bigrams(tokens))
                    listaTodosBigramas += bigramas
                    #tamBigramas = len(bigramas)
                    #for item in bigramas:
                    #    texto += '(' + item[0] + ' , ' + item[1] + ') | ' 
                    #self.gravarBigramasNoBD(rowId, texto, tamBigramas)
                    bigramas.clear()
            for item in listaTodosBigramas:
                cont = 0
                for bigrama in listaTodosBigramas:
                    if (item == bigrama):
                        cont += 1
                print(item, cont)
                # excluir para diminuir a lista
                while (item in listaTodosBigramas):
                    listaTodosBigramas.remove(item)

                # gravar contador 
                if tipoResposta.index(tipo) == 0: 
                    ws1.append([','.join(item), cont]) 
                else: 
                    ws2.append([','.join(item), cont]) 
            listaTodosBigramas.clear()
            wb.save(arq)

    def processaTrigramas(self):
        """ Percorre os textos e identifica os trigramas para geracao de planilha para analise
        """        
        # processa toda a tabela 
        wb = Workbook()
        ws1 = wb.create_sheet(title='trigramas-ANAMNESE')
        ws2 = wb.create_sheet(title='trigramas-EVOLUCAO')
        arq = 'trigramas.xlsx' 
        listaTodosTrigramas = [] 
        tipoResposta = ['ANAMNESE', 'EVOLUCAO']
        for tipo in tipoResposta:
            tabela = self.listarRespostas(tipo)
            for linha in tabela:
                rowId = linha[0]
                texto = ''
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    trigramas = list(nltk.trigrams(tokens))
                    listaTodosTrigramas += trigramas
                    tamtrigramas = len(trigramas)
                    for item in trigramas:
                       texto += '(' + item[0] + ' , ' + item[1] + ') | ' 
                    self.gravarTrigramasNoBD(rowId, texto, tamtrigramas)
                    trigramas.clear()
            for item in listaTodosTrigramas:
                cont = 0
                for trigrama in listaTodosTrigramas:
                    if (item == trigrama):
                        cont += 1
                print(item, cont)
                # excluir para diminuir a lista
                while (item in listaTodosTrigramas):
                    listaTodosTrigramas.remove(item)

                # gravar contador 
                if tipoResposta.index(tipo) == 0: 
                    ws1.append([','.join(item), cont]) 
                else: 
                    ws2.append([','.join(item), cont]) 
            listaTodosTrigramas.clear()
            wb.save(arq)
            print("Iteração salva")


if __name__ == "__main__":
    pass