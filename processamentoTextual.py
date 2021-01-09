import os
import constantes
import preProcessamentoTextual as pre 
import json 
import matplotlib.pyplot as plt
import nltk as nltk 
import pickle
from processamentoComBD import BD 
from os import path 
from wordcloud import WordCloud 
from openpyxl import Workbook 


def gerarNuvemDePalavrasPorArquivo(nomeArq, nomePng):
    """ Gera imagem com nuvem de palavras baseado em arquivo texto 
        Deve ser chamada depois do metodo separaStopWords() que gera o arquivo texto

    Args:
        nomeArq (str): Nome completo do arquivo (com path)
        nomePng (str): Nome do arquivo a ser salvo (sem path)
    """
    # get data directory (using getcwd() is needed to support running example in generated IPython notebook)
    d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()

    # Read the whole text.
    text = open(path.join(d, nomeArq)).read()

    # Generate a word cloud image
    wordcloud = WordCloud(width=1600, height=800, background_color="white", repeat=False, collocations=False)
    wordcloud.generate(text)
    wordcloud.to_file(constantes.PATH_RESULTADOS + nomePng)
    # Display the generated image:
    # the matplotlib way:
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")

    # lower max_font_size
    # wordcloud = WordCloud(max_font_size=40).generate(text)
    # plt.figure()
    # plt.imshow(wordcloud, interpolation="bilinear")
    # plt.axis("off")
    # plt.show()

def gerarNuvemDePalavrasPorDicionario(dicionario, nomePng):
    """ Gera imagem com nuvem de palavras baseado em estrutura de dados dicionario 

    Args:
        dicionario (dict): Estrutura de dados com formato chave:valor 
        nomePng (str): Nome do arquivo a ser salvo (sem path)
    """    
    #d = path.dirname(__file__) if "__file__" in locals() else os.getcwd()
    # Generate a word cloud image
    wordcloud = WordCloud(width=1600, height=800, background_color="white", repeat=False, collocations=False)
    wordcloud.generate_from_frequencies(dicionario)
    wordcloud.to_file(constantes.PATH_RESULTADOS + nomePng)
    # Display the generated image:
    # the matplotlib way:
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")

def processarSinaisSintomas():
    """ Lista e quantidade de sinais e sintomas 
        Saida: Sinais-sintomas.xlsx 
    """    
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'TODAS' 
    ws1 = wb.create_sheet(title='ANAMNESE') 
    ws2 = wb.create_sheet(title='EVOLUCAO') 
    arq = constantes.PATH_RESULTADOS + 'Sinais-sintomas.xlsx' 
    arqTodas = constantes.PATH_RESULTADOS + 'sinaisSintomas-ParaNuvem-Todas.txt'
    strTodas = ''
    arqAnamnese = constantes.PATH_RESULTADOS + 'sinaisSintomas-ParaNuvem-Anamnese.txt'
    strAnamnese = ''
    arqEvolucao = constantes.PATH_RESULTADOS + 'sinaisSintomas-ParaNuvem-Evolucao.txt'
    strEvolucao = ''
    colecao = {} 
    tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO'] 
    siglas = pre.carregarArquivoComoArray(constantes.ARQ_SINAIS_SINTOMAS) 
    for tipo in tipoResposta: 
        banco = BD(constantes.BD_SQL_RESPOSTAS)
        with banco:
            tabela = banco.listarRespostas(tipo) 
            for linha in tabela:
                if (linha[9]):
                    for sigla in siglas:
                        if sigla in linha[9]:
                            if sigla in colecao:
                                valor = colecao.get(sigla)
                                colecao[sigla] = valor + 1
                            else:
                                colecao[sigla] = 1
            for item in colecao:
                if tipoResposta.index(tipo) == 0:
                    ws0.append([item, colecao[item]]) 
                    strTodas += ' '+ pre.repeteString(item, colecao[item])
                elif tipoResposta.index(tipo) == 1:
                    ws1.append([item, colecao[item]]) 
                    strAnamnese += ' '+ pre.repeteString(item, colecao[item])
                elif tipoResposta.index(tipo) == 2:
                    ws2.append([item, colecao[item]]) 
                    strEvolucao += ' '+ pre.repeteString(item, colecao[item])
            colecao.clear()
            wb.save(arq)
    pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strTodas)).strip(), arqTodas)
    pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strAnamnese)).strip(), arqAnamnese)
    pre.gravarStringEmArquivo((pre.retirarExcessoDeEspacos(strEvolucao)).strip(), arqEvolucao)

def processarUniGramas():
    """ Percorre toda a tabela e identifica quais os tokens estao presentes em cada registro, 
        e realiza uma soma da presenca de cada token para uma visao geral dos termos mais usados
        Saida: 'UniGramas.xlsx' 
    """ 
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'TODAS' 
    ws1 = wb.create_sheet(title='ANAMNESE') 
    ws2 = wb.create_sheet(title='EVOLUCAO') 
    arq = constantes.PATH_RESULTADOS + 'UniGramas.xlsx' 
    colecao = {} 
    stopWords = pre.stopWords
    siglas = pre.carregarArquivoComoArray(constantes.ARQ_SIGLAS)
    sinaisSintomas = pre.carregarArquivoComoArray(constantes.ARQ_SINAIS_SINTOMAS)
    tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO']
    for tipo in tipoResposta:
        banco = BD(constantes.BD_SQL_RESPOSTAS)
        with banco:
            tabela = banco.listarRespostas(tipo)
            for linha in tabela:
                if (linha[9]):
                    tokens = pre.tokenizeString(linha[9])
                    for token in tokens:
                        if (not pre.possuiDigitoNumerico(token)) and (token not in stopWords) and (token not in siglas) and (token not in sinaisSintomas):
                            if token in colecao:
                                valor = colecao.get(token)
                                colecao[token] = valor + 1
                            else:
                                colecao[token] = 1 
        for chave, valor in colecao.items():
            if tipoResposta.index(tipo) == 0:
                ws0.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 1:
                ws1.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 2:
                ws2.append([chave, valor]) 
        colecao.clear()
        wb.save(arq)

def processaBigramas():
    """ Identifica e conta os bigramas. 
        Segmenta por Anamnese, Evolucao e Todos. 
    """        
    wb = Workbook()
    ws0 = wb.active
    ws0.title = 'bigramas-TODAS'
    ws1 = wb.create_sheet(title='bigramas-ANAMNESE')
    ws2 = wb.create_sheet(title='bigramas-EVOLUCAO')
    arq = constantes.PATH_RESULTADOS + 'bigramas.xlsx' 
    listaTodosBigramas = []
    tipoResposta = ['ANAMNESE', 'EVOLUCAO']
    colecao = {} 
    for tipo in tipoResposta:
        banco = BD(constantes.BD_SQL_RESPOSTAS) 
        with banco: 
            tabela = banco.listarRespostas(tipo) 
        for linha in tabela:
            if (linha[9]):
                tokens = pre.tokenizeString(linha[9])
                bigramas = list(nltk.bigrams(tokens))
                listaTodosBigramas += bigramas
                bigramas.clear()
        for item in listaTodosBigramas:
            for bigrama in listaTodosBigramas:
                if (item == bigrama):
                    if item in colecao:
                        valor = colecao.get(item)
                        colecao[item] = valor + 1 
                    else:
                        colecao[item] = 1
            # excluir para diminuir a lista
            while (item in listaTodosBigramas):
                listaTodosBigramas.remove(item)
            # gravar contador 
            if tipoResposta.index(tipo) == 0: 
                ws1.append([','.join(item), colecao.get(item)]) 
            else: 
                ws2.append([','.join(item), colecao.get(item)]) 
        listaTodosBigramas.clear()
        wb.save(arq)
        if tipoResposta.index(tipo) == 0:
            arqJson = constantes.PATH_RESULTADOS + "bigramas-dict-Anamnese.pickle" 
        else:
            arqJson = constantes.PATH_RESULTADOS + "bigramas-dict-Evolucao.pickle" 
        with open(arqJson, 'wb') as f:
            pickle.dump(colecao, f)
        colecao.clear() 

def carregaDicionarioComMinimoDeItens(arqDictPickle, numValorMinimo):
    """ Recebe um arquivo de dicionario com lista, valor no formato pickle e exclui os itens abaixo do minimo indicado

    Args:
        arqDictPickle (pickle): Arquivo de dicionario em formato pickle
        numValorMinimo (int): Numero minimo a considerar como valor do item no dicionario

    Returns:
        dict: Dicionario em memoria considerando valor minimo por item
    """    
    with open(arqDictPickle, 'rb') as f:
        dicionario = pickle.loads(f.read())
        copia = dicionario.copy()
        for chave, valor in copia.items():
            if (int(valor) < int(numValorMinimo)):
                dicionario.pop(chave)
        resposta = {}
        for itens, valor in dicionario.items():
            texto = ""
            for item in itens:
                texto = texto + " " + item 
            resposta[texto.strip()] = valor
        return resposta

def processaTrigramas():
    """ Percorre os textos e identifica os trigramas para geracao de planilha para analise
        Segmenta por Anamnese e Evolucao 
        Grava como saida 
    """        
    wb = Workbook()
    ws1 = wb.create_sheet(title='trigramas-ANAMNESE')
    ws2 = wb.create_sheet(title='trigramas-EVOLUCAO')
    arq = constantes.PATH_RESULTADOS + 'trigramas.xlsx' 
    listaTodosTrigramas = [] 
    tipoResposta = ['ANAMNESE', 'EVOLUCAO']
    colecao = {} 
    for tipo in tipoResposta:
        banco = BD(constantes.BD_SQL_RESPOSTAS) 
        with banco: 
            tabela = banco.listarRespostas(tipo) 
        for linha in tabela:
            if (linha[9]):
                tokens = pre.tokenizeString(linha[9])
                trigramas = list(nltk.trigrams(tokens))
                listaTodosTrigramas += trigramas
                trigramas.clear()
        for item in listaTodosTrigramas:
            for trigrama in listaTodosTrigramas:
                if (item == trigrama):
                    if (item in colecao):
                        valor = colecao.get(item)
                        colecao[item] = valor + 1 
                    else: 
                        colecao[item] = 1
            # excluir para diminuir a lista
            while (item in listaTodosTrigramas):
                print(item)
                listaTodosTrigramas.remove(item)
            # gravar contador 
            if tipoResposta.index(tipo) == 0: 
                ws1.append([','.join(item), colecao.get(item)]) 
            else: 
                ws2.append([','.join(item), colecao.get(item)]) 
        listaTodosTrigramas.clear()
        wb.save(arq)
        if tipoResposta.index(tipo) == 0:
            arqJson = constantes.PATH_RESULTADOS + "trigramas-dict-Anamnese.pickle"
        else: 
            arqJson = constantes.PATH_RESULTADOS + "trigramas-dict-Evolucao.pickle"
        with open (arqJson, 'wb') as f:
            pickle.dump(colecao,f)
        colecao.clear()

def processarTerminologia(arqTerminologia, nomeArqSaida): 
    """ Verifica no banco de dados se ha termos da terminologia em cada documento
        baseado na lista de terminologia previamente selecionada 
    Args:
        arqTerminologia (str): Arquivo completo (com path) de qual terminologia usar (vide constantes)
        nomeArqSaida ([type]): Nome do arquivo de saida a ser gravado no PATH_RESULTADOS 
    """
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'TODAS' 
    ws1 = wb.create_sheet(title='ANAMNESE') 
    ws2 = wb.create_sheet(title='EVOLUCAO') 
    arq = constantes.PATH_RESULTADOS + nomeArqSaida 
    colecao = {} 
    terminologia = pre.carregarArquivoComoArray(arqTerminologia) 
    tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO'] 
    for tipo in tipoResposta: 
        banco = BD(constantes.BD_SQL_RESPOSTAS) 
        with banco: 
            tabela = banco.listarRespostas(tipo) 
            for linha in tabela: 
                if (linha[9]): 
                    for exp in terminologia: 
                        if (exp in linha[9]): 
                            if exp in colecao: 
                                valor = colecao.get(exp) 
                                colecao[exp] = valor + 1 
                            else: 
                                colecao[exp] = 1 
        for chave, valor in colecao.items():
            if tipoResposta.index(tipo) == 0:
                ws0.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 1:
                ws1.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 2:
                ws2.append([chave, valor]) 
        colecao.clear()
        wb.save(arq)

def processarTerminologiaComNGramas(arqTerminologia, nomeArqSaida, quantNGramas):
    """ Dados um conjunto terminologico, verifica cada termo com seus proximos nGramas

    Args:
        arqTerminologia (str): Arquivo completo (com path) de qual terminologia usar (vide constantes)
        nomeArqSaida (str): Nome do arquivo de saida a ser gravado no PATH_RESULTADOS 
        quantNGramas (int): Numero de nGramas a serem considerados
    """    
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'TODAS' 
    ws1 = wb.create_sheet(title='ANAMNESE') 
    ws2 = wb.create_sheet(title='EVOLUCAO') 
    arq = constantes.PATH_RESULTADOS + nomeArqSaida 
    colecao = {} 
    terminologia = pre.carregarArquivoComoArray(arqTerminologia) 
    tipoResposta = ['TODAS', 'ANAMNESE', 'EVOLUCAO'] 
    for tipo in tipoResposta: 
        banco = BD(constantes.BD_SQL_RESPOSTAS) 
        with banco: 
            tabela = banco.listarRespostas(tipo) 
            for linha in tabela: 
                if (linha[9]): 
                    for termo in terminologia: 
                        if (termo in linha[9]): 
                            nGram = ""
                            tokens = pre.tokenizeString(linha[9])
                            try:
                                indIni = tokens.index(termo)
                            except:
                                indIni = -1
                            if indIni > -1:
                                indFim = indIni + quantNGramas
                                if (indFim > len(tokens)):
                                    indFim = len(tokens)
                                for i in range(indIni, indFim):
                                    nGram = str(nGram) + " " + tokens[i]
                                if nGram in colecao: 
                                    valor = colecao.get(nGram) 
                                    colecao[nGram] = valor + 1 
                                else: 
                                    colecao[nGram] = 1 
        for chave, valor in colecao.items():
            if tipoResposta.index(tipo) == 0:
                ws0.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 1:
                ws1.append([chave, valor]) 
            elif tipoResposta.index(tipo) == 2:
                ws2.append([chave, valor]) 
        colecao.clear()
        wb.save(arq)

